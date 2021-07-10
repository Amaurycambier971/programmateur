#IMPORTATION REQUISE
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os
import time

os.chdir('/Users/amaurycambier/Documents/Bamy2')
borderStyle = openpyxl.styles.Side(style = "thin")

def tri(elem, colone) :
    i=2
    j=2
    while ws[f"{colone}{i}"].value != None :
        if elem in str(ws[f"{colone}{i}"].value) :
            for k in "ABCDEFGHIJKLMNOPQRSTUVWXY" :
                ws1[f"{k}{j}"].value = ws[f"{k}{i}"].value
                ws1[f"Z{j}"]= 2021
            j+=1
        i+=1
#DETERMINER LES INTRANTS !

NomF = input("Quel est le nom du fichier source ? (rentrez exactement le nom, .xlsx inclus \nsaisir le nom du fichier source :")
print("\n")

triENC = input("Quel est la colone du type de pièce ? (ENC) \nsaisisez la lettre : ")
while triENC not in "ABCDEFGHIJKLMNOPQRSTUVWXYZ" :
    triENC = input("Quel est la colone du type de pièce ? (ENC) \nsaisisez la lettre : ")

print("\n")
print("vous allez maintenant devoir saisir les noms des colones utile pour le TCD (les noms, pas les lettres !")
annee = input("saisir le nom de la colone donnant les années : ")
NumP = input("saisir le nom de la colone donnant les Numéros de pièce : ")
Tier = input("saisir le nom de la colone donnant les numéros de tier : ")
Debit = input("saisir le nom de la colone donnant les Débit : ")
credit = input("saisir le nom de la colone donnant les crédit : ")

#ECRITURE3.xlsx
#S
#ANNEE
#Num pièce




start = time.time()



#ON PRELEVE LE FICHIER DONT ON A BESOIN
ecriture = pd.read_excel(NomF, "ECR")
#ON VA MAINTENANT CREER NOTRE FICHIER EXCEL
file = 'Fractmt rglt sur client.xlsx'
writer = pd.ExcelWriter(file, engine='xlsxwriter')
ecriture.to_excel(writer, sheet_name='compte à compte', index=False)
writer.save()

#ON AJOUTE A NOTRE FICHIER UN SECOND ONGLET


wb = openpyxl.load_workbook(filename=file)
ws = wb['compte à compte']
ws1 = wb.create_sheet("compte a compte 2")
ws1.title = "compte a compte 2"







# IL FAUDRAS SUPPRIMER CA SI IL Y A DEJA LA COLONNE "ANNEE"

#ON COPIE COLLE LA PREMIER LIGNE DU ONGLET 1 AU 2
for i in "ABCDEFGHIJKLMNOPQRSTUVWXY" :
    ws1[f"{i}1"].value = ws[f"{i}1"].value
ws1["Z1"].value = "ANNEE"



#COPIER COLLER = TRIE (pas de probleme pour la colone, c'est une variable)

tri("ENC",triENC)

wb.save(file)


# ON VA CREER LA PAGE TCD
df = pd.read_excel(file, sheet_name="compte a compte 2")

pivot = pd.pivot_table(df, index=[annee, NumP,Tier], values=[Debit,credit], aggfunc="sum")
writer = pd.ExcelWriter(file, engine='openpyxl', mode="a")
pivot = pivot.reindex(columns=[Debit,credit])
pivot.to_excel(writer, sheet_name='TCD')



writer.save()

#ON VA COLORIER LES CELLULES POSANT PROBLEME
wb = openpyxl.load_workbook(filename=file)
ws = wb['TCD']

i=2

# idée : pour les TCD type excel on peut comparer le type : int, str pour faire la démarcation

while ws[f"C{i}"].value != None :
    if ws[f"B{i}"].value == None :
        ws[f"C{i}"].fill = PatternFill("solid", fgColor="FF0000")
        ws[f"C{i-1}"].fill = PatternFill("solid", fgColor="FF0000")
    i+=1

wb.save(file)

#ON CREER UN TCD REMPLIE UNIQUEMENT AVEC LES CASES A PROBLEME

wb = openpyxl.load_workbook(filename=file)
ws = wb["TCD"]
ws1 = wb.create_sheet("TCD Final")

# TCD FINAL
for i in "ABCDE" :
    ws1[f"{i}1"].value = ws[f"{i}1"].value
    ws1[f"{i}1"].font = openpyxl.styles.Font(bold="True")
    ws1[f"{i}1"].border = openpyxl.styles.Border(left = borderStyle, right=borderStyle, top =borderStyle, bottom = borderStyle)
i=2
while ws[f"C{i}"].value != None :
    ws1[f"A{i}"].value = ws[f"A{i}"].value
    ws1[f"A{i}"].font = openpyxl.styles.Font(bold="True")
    if ws[f"A{i}"].value == None :
        ws1[f"A{i}"].fill = PatternFill("solid", fgColor="FFFFFF")
    i+=1


#ALGORITHME DE LA SELECTION DES VALEURS


i=2
j=2
k=2
while ws[f"C{i}"].value != None :
    if ws[f"B{i}"].value != None and ws[f"B{i+1}"].value == None:
        ws1[f"B{j}"].value = ws[f"B{i}"].value
        ws1[f"B{j}"].font = openpyxl.styles.Font(bold="True")
        ws1[f"C{j}"].value = ws[f"C{i}"].value
        ws1[f"C{j}"].font = openpyxl.styles.Font(bold="True")
        ws1[f"C{j}"].border = openpyxl.styles.Border(left = borderStyle, right=borderStyle, top =borderStyle, bottom = borderStyle)

        ws1[f"D{j}"].value = ws[f"D{i}"].value
        ws1[f"E{j}"].value = ws[f"E{i}"].value


        j+=1

    if ws[f"B{i}"].value == None:
        ws1[f"B{j}"].value = ws[f"B{i}"].value
        ws1[f"B{j}"].fill = PatternFill("solid", fgColor="FFFFFF")


        ws1[f"C{j}"].value = ws[f"C{i}"].value
        ws1[f"C{j}"].fill = PatternFill("solid", fgColor="FFFFFF")
        ws1[f"C{j}"].font = openpyxl.styles.Font(bold="True")
        ws1[f"C{j}"].border = openpyxl.styles.Border(left = borderStyle, right=borderStyle, top =borderStyle, bottom = borderStyle)

        ws1[f"D{j}"].value = ws[f"D{i}"].value
        ws1[f"E{j}"].value = ws[f"E{i}"].value
        j+=1

    ws1[f"B{i}"].fill = PatternFill("solid", fgColor="FFFFFF")

    i+=1


# FUSION DES CELLULES POUR MISE EN PAGE + BORDURE

i = 2
p1=2
#fusion de la colone A
while ws1[f"C{i}"].value != None :
    if ws1[f"A{i}"].value != None and i>3:
        ws1.merge_cells(f"A{p1}:A{i}")
        p1 = i
    i+=1
i-=1
ws1.merge_cells(f"A{p1}:A{i}")
#fusion de la colone B
i=2
while ws1[f"C{i}"].value != None :
    if ws1[f"B{i}"].value != None and i>3:
        ws1.merge_cells(f"B{p1}:B{i-1}")
        p1 = i
    i+=1

i=2
for j in "B" :
    while ws1[f"C{i}"].value != None :
        ws1[f"{j}{i}"].border = openpyxl.styles.Border(left = borderStyle, right=borderStyle, top =borderStyle, bottom = borderStyle)
        ws1[f"{j}{i}"].alignment = openpyxl.styles.Alignment(vertical="top",horizontal="center")
        i+=1
i=2
for j in "A" :
    while ws1[f"C{i}"].value != None :
        ws1[f"{j}{i}"].border = openpyxl.styles.Border(left = borderStyle, right=borderStyle, top =borderStyle, bottom = borderStyle)
        ws1[f"{j}{i}"].alignment = openpyxl.styles.Alignment(vertical="top",horizontal="center")
        i+=1



wb.save(file)

end = time.time()
temps = end-start

print("Le programme à mis ", round(temps,2), " secondes à s'effectuer")

